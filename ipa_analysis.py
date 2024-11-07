"""
This script processes data for pathway analysis, including splitting categories,
extracting top subcategories, and creating visualization plots.

Requirements:
- EF1_RUV_upregulated_disease.csv (your IPA output files)
- Final_combined_unique_categories_pathways.csv
- DE_results_RUVs_normalized_gene_symbol_EF1.csv (your DEG files)

Usage:
Run this script in the command line using:
    python data_analysis_pipeline.py
"""

# Code starts here
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import seaborn as sns
import os
from openpyxl import Workbook


# Set your input filename and derive the prefix
input_filename = "IPA_FC1.1_EF1_downregulated.csv"
DEG=pd.read_csv("DE_results_RUVs_normalized_gene_symbol_EF1.csv")
pathway = pd.read_csv("Final_combined_unique_categories_pathways.csv")

input_prefix = os.path.splitext(input_filename)[0]  
output_dir = input_prefix
os.makedirs(output_dir, exist_ok=True)

# Read the main data file
df = pd.read_csv(input_filename)
df['Categories'] = df['Categories'].str.replace('DNA Replication, Recombination, and Repair', 'DNA Replication Recombination and Repair')

# Read the pathway file
pathway = pd.read_csv("Final_combined_unique_categories_pathways.csv")


# prompt: split Categories column by commas and save them into new columns 1,2,3,4,5,6.....

categories_split = df['Categories'].str.split(',', expand=True)
categories_split.columns = ['Category_' + str(i) for i in range(1, categories_split.shape[1] + 1)]
df = pd.concat([df, categories_split], axis=1)

# Map pathways to categories without replacing original values
for index, row in pathway.iterrows():
    category = row['Categories']
    pathway_value = row['Pathways']
    for col in df.columns:
        if col.startswith('Category_'):
            df.loc[df[col] == category, col] = pathway_value

# prompt: Combine the duplicated value in Category_1 to Category_11, and only save one value, and remove the empty column

category_columns = [col for col in df.columns if col.startswith('Category_')]

# Iterate over rows to combine unique categories
for index, row in df.iterrows():
    # Collect unique, non-null categories for each row
    unique_categories = set(row[col] for col in category_columns if pd.notna(row[col]))

    # Assign unique categories back to the columns
    for i, col in enumerate(category_columns):
        df.loc[index, col] = list(unique_categories)[i] if i < len(unique_categories) else None

# Drop any columns that are completely empty
df = df.dropna(axis=1, how='all')

# Split Categories column by commas and save to new subcategory columns
categories_split = df['Categories'].str.split(',', expand=True)
categories_split.columns = ['subcategory_' + str(i) for i in range(1, categories_split.shape[1] + 1)]
df = pd.concat([df, categories_split], axis=1)


# Assuming your DataFrame is named 'df' and the category columns are 'Category_1', 'Category_2', and 'Category_3'

# Assuming your DataFrame is named 'df' and the category columns are 'Category_1', 'Category_2', and 'Category_3'

# Define the list of category columns
category_columns = [col for col in df.columns if col.startswith('Category_')]

# Use these columns for filtering based on the specified values
df_diseases = df[df[category_columns].apply(lambda row: 'Diseases_and_Disorders' in row.values, axis=1)]
df_physiological = df[df[category_columns].apply(lambda row: 'Physiological_System_Development_and_Function' in row.values, axis=1)]
df_molecular = df[df[category_columns].apply(lambda row: 'Molecular_and_Cellular_Functions' in row.values, axis=1)]




# Display the counts of rows in each new DataFrame
print(f"Number of rows in df_diseases: {len(df_diseases)}")
print(f"Number of rows in df_physiological: {len(df_physiological)}")
print(f"Number of rows in df_molecular: {len(df_molecular)}")


# prompt: extract the corresponding Values for the pathway["Pathways"] and create 3 list disease, physiological,molecular from the pathway df

disease_list = pathway[pathway['Pathways'] == 'Diseases_and_Disorders']['Categories'].tolist()
physiological_list = pathway[pathway['Pathways'] == 'Physiological_System_Development_and_Function']['Categories'].tolist()
molecular_list = pathway[pathway['Pathways'] == 'Molecular_and_Cellular_Functions']['Categories'].tolist()

# Display the lists to verify
print("Disease Categories:", disease_list)
print("Physiological Categories:", physiological_list)
print("Molecular Categories:", molecular_list)

# prompt: Look through the subcategory_* columns, and find the unique values and extract the top 10 values in the subcategory_* that has the smallest "B-H p-value" and do it for the 3 dfs(df_diseases,df_physiological,df_molecular). each dfs value need to be disease_list or physiological_list or molecular_list

def extract_top_10_subcategories(df, pathway_list):
    """Extracts the top 10 subcategories with the smallest 'B-H p-value' for a given DataFrame and pathway list.

    Args:
        df: The DataFrame containing the data.
        pathway_list: A list of pathway categories to consider.

    Returns:
        A list of the top 10 subcategories with the smallest 'B-H p-value'.
    """
    subcategory_cols = [col for col in df.columns if col.startswith('subcategory_')]
    subcategory_min_pvalues = {}

    for subcategory_col in subcategory_cols:
        # Filter the DataFrame for rows where the subcategory is in the given pathway list
        df_filtered = df[df[subcategory_col].isin(pathway_list)]
        if not df_filtered.empty:
            # Sort the filtered DataFrame by 'B-H p-value' in ascending order
            df_sorted = df_filtered.sort_values('B-H p-value')
            # Iterate over the rows to capture the minimum p-value for each unique subcategory
            for subcategory, p_value in zip(df_sorted[subcategory_col], df_sorted['B-H p-value']):
                if subcategory not in subcategory_min_pvalues or p_value < subcategory_min_pvalues[subcategory]:
                    subcategory_min_pvalues[subcategory] = p_value

    # Sort subcategories by the smallest p-value and select the top 10
    sorted_subcategories = sorted(subcategory_min_pvalues.items(), key=lambda item: item[1])[:10]
    top_10_subcategories = [subcategory for subcategory, _ in sorted_subcategories]

    return top_10_subcategories


# Example usage with the pathway lists
disease_top_10 = extract_top_10_subcategories(df_diseases, disease_list)
physiological_top_10 = extract_top_10_subcategories(df_physiological, physiological_list)
molecular_top_10 = extract_top_10_subcategories(df_molecular, molecular_list)

# Display the results
print("Top 10 Disease Subcategories:", disease_top_10)
print("Top 10 Physiological Subcategories:", physiological_top_10)
print("Top 10 Molecular Subcategories:", molecular_top_10)

disease_top_10

# prompt: I want to extract separate dfs from the 3 dfs(df_diseases,df_physiological,df_molecular) with the values in the "subcategory_*" column that has top_10_diseases, top_10_physiological,top_10_molecular (one dataframe with only one values in the list (top_10_diseases, top_10_physiological,top_10_molecular ) , so i will have 30 dfs. (please name the dfs with the value names in each dfs)

# Assuming you have df_diseases, df_physiological, and df_molecular DataFrames

def create_dfs_for_subcategories(df, subcategory_list):
  """Creates separate DataFrames for each subcategory in the list."""
  dfs = {}
  for subcategory in subcategory_list:
    dfs[subcategory] = df[df.apply(lambda row: subcategory in row.values, axis=1)]
  return dfs

# Create separate DataFrames for each subcategory in the top 10 lists
dfs_diseases = create_dfs_for_subcategories(df_diseases, disease_top_10)
dfs_physiological = create_dfs_for_subcategories(df_physiological, physiological_top_10)
dfs_molecular = create_dfs_for_subcategories(df_molecular, molecular_top_10)

# Combine the dictionaries into one
all_dfs = {**dfs_diseases, **dfs_physiological, **dfs_molecular}

# You can access individual DataFrames using the subcategory name as the key
# For example, to access the DataFrame for 'Cancer' (if it's in top_10_diseases):
#df_cardio = all_dfs['Cardiovascular System Development and Function']

# Create separate DataFrames based on subcategories, only if they exist in all_dfs
for subcategory in disease_top_10:
    if subcategory in all_dfs:
        df_name = "df_" + subcategory.replace(" ", "_")
        globals()[df_name] = all_dfs[subcategory]

for subcategory in physiological_top_10:
    if subcategory in all_dfs:
        df_name = "df_" + subcategory.replace(" ", "_")
        globals()[df_name] = all_dfs[subcategory]

for subcategory in molecular_top_10:
    if subcategory in all_dfs:
        df_name = "df_" + subcategory.replace(" ", "_")
        globals()[df_name] = all_dfs[subcategory]

# Print the contents of each list to verify the subcategory names
print("Top 10 Subcategories (Diseases):", disease_top_10)
print("Top 10 Subcategories (Physiological):", physiological_top_10)
print("Top 10 Subcategories (Molecular):", molecular_top_10)

# Print all keys in all_dfs to confirm they match the subcategory names in the lists
print("\nKeys in all_dfs:", list(all_dfs.keys()))

df_names = [name for name in globals() if name.startswith("df_") and isinstance(globals()[name], pd.DataFrame)]
df_names, len(df_names)



# prompt: provide me with an excel sheet with the subsheets of the globals dfs (Not including 'df_diseases', 'df_physiological', 'df_molecular' ) subsheets
# Assuming you have df_names defined as a list of DataFrame names

# Create a new Excel workbook
workbook = Workbook()

# Iterate through DataFrame names
for df_name in df_names:
    if df_name not in ['df_diseases', 'df_physiological', 'df_molecular']:
        try:
            # Get the DataFrame from globals()
            df = globals()[df_name]

            # Create a new sheet in the workbook with the DataFrame name (truncate if necessary)
            sheet_name = df_name[:31]  # Excel sheet names are limited to 31 characters
            sheet = workbook.create_sheet(title=sheet_name)

            # Write column headers
            for col_idx, column_name in enumerate(df.columns, start=1):
                sheet.cell(row=1, column=col_idx, value=column_name)

            # Write DataFrame values, converting scientific notation to strings
            for row_idx, row in enumerate(df.values, start=2):
                for col_idx, value in enumerate(row, start=1):
                    # Convert scientific notation to string if it's a large or small float
                    if isinstance(value, float) and (value > 1e10 or value < 1e-10):
                        sheet.cell(row=row_idx, column=col_idx, value=str(value))
                    else:
                        sheet.cell(row=row_idx, column=col_idx, value=value)

        except KeyError:
            # Handle cases where a DataFrame with the given name doesn't exist
            print(f"DataFrame '{df_name}' not found.")

# Remove the default sheet created by Workbook()
if 'Sheet' in workbook.sheetnames:
    workbook.remove(workbook['Sheet'])

# Save the workbook to a file
workbook.save(os.path.join(output_dir, f"globals_dfs_subsheets_{input_prefix}_top10pathways.xlsx"))

import pandas as pd

def get_gene_list(df):
    """Extracts a master gene list from a DataFrame's Molecules column."""
    if 'Molecules' not in df.columns:
        return []  # Handle cases where the column doesn't exist

    gene_list = set()  # Use a set to avoid duplicates
    for genes in df['Molecules'].dropna():  # Ignore rows with NaN in 'Molecules'
        for gene in genes.split(','):
            gene_list.add(gene.strip())  # Remove leading/trailing whitespace
    return list(gene_list)


# List of DataFrame names in globals() that start with "df_" but are not the main category DataFrames
df_names = [name for name in globals() if name.startswith("df_")
            and name not in ('df_diseases', 'df_physiological', 'df_molecular')
            and isinstance(globals()[name], pd.DataFrame)]

# List to store each DataFrame's information as a row for the final CSV
all_data = []

# Loop through each DataFrame, extract information, and add it to the list
for df_name in df_names:
    df = globals()[df_name]

    # Extract the pathway name from the DataFrame name
    pathway = df_name.replace("df_", "").replace("_", " ")

    # Determine the subcategory type based on the top 10 lists
    if pathway in disease_top_10:
        subcategory_type = "Diseases"
    elif pathway in physiological_top_10:
        subcategory_type = "Physiological"
    elif pathway in molecular_top_10:
        subcategory_type = "Molecular"
    else:
        subcategory_type = "Other"

    # Get the smallest "B-H p-value" in the DataFrame
    smallest_p_value = df['B-H p-value'].min() if 'B-H p-value' in df.columns else None

    # Get the master gene list
    gene_list = get_gene_list(df)
    gene_count = len(gene_list)  # Count the number of unique genes

    # Prepare the row for the CSV file
    row = {
        'Pathway': pathway,
        'Subcategory Type': subcategory_type,  # Add the new subcategory type column
        'Smallest B-H p-value': smallest_p_value,
        'Number of Genes': gene_count
    }

    # Add genes as separate columns (starting from column 5)
    for i, gene in enumerate(gene_list, start=1):
        row[f'Gene_{i}'] = gene

    # Append the row dictionary to all_data
    all_data.append(row)

# Convert all_data to a DataFrame and save to a single CSV file
final_df = pd.DataFrame(all_data)
final_df.to_csv(os.path.join(output_dir, f"all_dfs_summary_{input_prefix}_top10pathways.csv"), index=False)


final_disease=final_df.loc[final_df["Subcategory Type"] == "Diseases"]
final_physiological=final_df.loc[final_df["Subcategory Type"] == "Physiological"]
final_molecular=final_df.loc[final_df["Subcategory Type"] == "Molecular"]

final_disease = final_disease.sort_values(by="Smallest B-H p-value", ascending=False)
final_physiological = final_physiological.sort_values(by="Smallest B-H p-value", ascending=False)
final_molecular = final_molecular.sort_values(by="Smallest B-H p-value", ascending=False)

# prompt: plot barh plot based on log padj top_30_pvalues"B-H p-value" and the y axis is "Diseases or Functions Annotation" display the top 10 pathways

import matplotlib.pyplot as plt
import numpy as np

# Assuming 'top_30_pvalues' is your DataFrame
plt.figure(figsize=(7,6))
plt.barh(final_disease['Pathway'], -np.log10(final_disease['Smallest B-H p-value']), color="lightblue")
plt.xlabel('-log10(B-H p-value)')
plt.ylabel('Diseases and Disorders')
plt.title('Top 10 Pathways by -log10(B-H p-value)')
plt.tight_layout()
plt.savefig(os.path.join(output_dir, f"Top_10_disease_{input_prefix}.png"), format="png", dpi=300)

# prompt: plot barh plot based on log padj top_30_pvalues"B-H p-value" and the y axis is "Diseases or Functions Annotation" display the top 10 pathways

import matplotlib.pyplot as plt
import numpy as np

# Assuming 'final_physiological' is your DataFrame
plt.figure(figsize=(8, 6))
plt.barh(final_physiological['Pathway'], -np.log10(final_physiological['Smallest B-H p-value']), color="teal")
plt.xlabel('-log10(B-H p-value)')
plt.ylabel('Physiological System Development and Function')
plt.title('Top 10 Pathways by -log10(B-H p-value)')
plt.tight_layout()
plt.savefig(os.path.join(output_dir,f"Top_10_physiological_{input_prefix}.png"), format="png", dpi=300)

# prompt: plot barh plot based on log padj top_30_pvalues"B-H p-value" and the y axis is "Diseases or Functions Annotation" display the top 10 pathways

import matplotlib.pyplot as plt
import numpy as np

# Assuming 'final_physiological' is your DataFrame
plt.figure(figsize=(7, 6))
plt.barh(final_molecular['Pathway'], -np.log10(final_molecular['Smallest B-H p-value']), color="lightgreen")
plt.xlabel('-log10(B-H p-value)')
plt.ylabel('Molecular and Cellular Functions')
plt.title('Top 10 Pathways by -log10(B-H p-value)')
plt.tight_layout()
plt.savefig(os.path.join(output_dir,f"Top_10_molecular_{input_prefix}.png"), format="png", dpi=300)

# prompt: add a -logpadh column by -log(B-H p-value)

final_disease['-logpadj'] = -np.log10(final_disease['Smallest B-H p-value'])
final_physiological['-logpadj'] = -np.log10(final_physiological['Smallest B-H p-value'])
final_molecular['-logpadj'] = -np.log10(final_molecular['Smallest B-H p-value'])

final_disease

final_disease["GeneRatio"] = final_disease["Number of Genes"] / 326
final_physiological["GeneRatio"] = final_physiological["Number of Genes"] / 326
final_molecular["GeneRatio"] = final_molecular["Number of Genes"] / 326

import seaborn as sns

plt.figure(figsize=(10, 6))
scatter = sns.scatterplot(
    data=final_disease,
    x="GeneRatio",  # Use 'GeneRatio' as the x-axis
    y="Pathway",  # Use 'Diseases or Functions Annotation' as the y-axis
    hue="-logpadj",  # Color by 'B-H p-value'
    size="GeneRatio",  # Size by 'GeneRatio'
    sizes=(20, 200),  # Size range for the points
    palette="coolwarm",  # Color gradient from 'coolwarm'
    legend="brief"
)

# Customize plot appearance
plt.title(f"{input_prefix} Diseases and Disorders")
plt.xlabel("Gene ratio")
plt.ylabel("")
plt.legend(title="B-H p-value & GeneRatio", bbox_to_anchor=(1, 1))
plt.tight_layout()
plt.savefig(os.path.join(output_dir,f"{input_prefix}_Diseases_dotplots.png"), format="png", dpi=300)

plt.figure(figsize=(11.5, 6))
scatter = sns.scatterplot(
    data=final_physiological,
    x="GeneRatio",  # Use 'GeneRatio' as the x-axis
    y="Pathway",  # Use 'Diseases or Functions Annotation' as the y-axis
    hue="-logpadj",  # Color by 'B-H p-value'
    size="GeneRatio",  # Size by 'GeneRatio'
    sizes=(20, 200),  # Size range for the points
    palette="coolwarm",  # Color gradient from 'coolwarm'
    legend="brief"
)

# Customize plot appearance
plt.title(f"{input_prefix} Physiological System Development and Function")
plt.xlabel("Gene ratio")
plt.ylabel("")
plt.legend(title="B-H p-value & GeneRatio", bbox_to_anchor=(1, 1))
plt.tight_layout()
plt.savefig(os.path.join(output_dir,f"{input_prefix}_physiological_dotplots.png"), format="png", dpi=300)

plt.figure(figsize=(10, 6))
scatter = sns.scatterplot(
    data=final_molecular,
    x="GeneRatio",  # Use 'GeneRatio' as the x-axis
    y="Pathway",  # Use 'Diseases or Functions Annotation' as the y-axis
    hue="-logpadj",  # Color by 'B-H p-value'
    size="GeneRatio",  # Size by 'GeneRatio'
    sizes=(20, 200),  # Size range for the points
    palette="coolwarm",  # Color gradient from 'coolwarm'
    legend="brief"
)

# Customize plot appearance
plt.title(f"{input_prefix} Molecular and Cellular Functions")
plt.xlabel("Gene ratio")
plt.ylabel("")
plt.legend(title="B-H p-value & GeneRatio", bbox_to_anchor=(1, 1))
plt.tight_layout()
plt.savefig(os.path.join(output_dir,f"{input_prefix}_molecular_dotplots.png"), format="png", dpi=300)

# Replace ', ' with a tab in top_10_pathways["Molecules"]
final_molecular


gmt_dir = os.path.join(input_prefix, f"{input_prefix}_GSEA_gmt_top10")
# Create the directory if it doesn't exist
os.makedirs(gmt_dir, exist_ok=True)

def create_gmt_file(df):
    """Creates separate .gmt files from a DataFrame for each pathway."""
    for index, row in df.iterrows():
        pathway_name = row['Pathway']

        # Extract genes from columns that start with 'Gene_' and are not NaN
        genes = [row[gene] for gene in df.columns if gene.startswith('Gene_') and pd.notna(row[gene])]

        # Construct the gene list string
        gene_list_string = '\t'.join(genes)

        # Write to a .gmt file with the pathway name as the filename
        if gene_list_string:  # Only write if there are genes
            filename = os.path.join(gmt_dir, f"{input_prefix}_{pathway_name.replace(' ', '_')}.gmt")
            with open(filename, 'w') as f:
                f.write(f"{pathway_name}\tNA\t{gene_list_string}\n")
            print(f"Created {filename}")

# Apply to each DataFrame
for df in [final_disease, final_physiological, final_molecular]:
    create_gmt_file(df)


# Create a new Excel workbook
workbook = Workbook()

# Get column headers from the DEG DataFrame
deg_columns = list(DEG.columns)

# Iterate through the rows of 'final_disease'
for index, row in final_disease.iterrows():
    pathway = row['Pathway']

    # Create a new sheet for each pathway with the title
    sheet = workbook.create_sheet(title=pathway[:31])  # Limit to 31 characters (Excel sheet name limit)

    # Write DEG column headers to the first row
    for col_idx, header in enumerate(deg_columns, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    # Get the gene columns from 'final_disease' row
    gene_columns = [col for col in final_disease.columns if col.startswith('Gene_')]
    genes = [row[col] for col in gene_columns if pd.notna(row[col])]

    # Write gene data to the sheet by checking if each gene is in DEG
    row_index = 2  # Start after header row
    for gene in genes:
        if gene in DEG['Name'].values:
            matching_rows = DEG.loc[DEG['Name'] == gene].values.tolist()
            for match_row in matching_rows:
                # Write matching row data from DEG
                for col_idx, value in enumerate(match_row, start=1):
                    sheet.cell(row=row_index, column=col_idx, value=value)
                row_index += 1

# Remove the default sheet created by Workbook() if it exists
if 'Sheet' in workbook.sheetnames:
    workbook.remove(workbook['Sheet'])

# Save the workbook to a file
workbook.save(os.path.join(output_dir, f"{input_prefix}_final_disease_genes_in_DEG_top10pathways.xlsx"))


# Create a new Excel workbook
workbook = Workbook()

# Get column headers from the DEG DataFrame
deg_columns = list(DEG.columns)

# Iterate through the rows of 'final_disease'
for index, row in final_physiological.iterrows():
    pathway = row['Pathway']

    # Create a new sheet for each pathway with the title
    sheet = workbook.create_sheet(title=pathway[:31])  # Limit to 31 characters (Excel sheet name limit)

    # Write DEG column headers to the first row
    for col_idx, header in enumerate(deg_columns, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    # Get the gene columns from 'final_disease' row
    gene_columns = [col for col in final_physiological.columns if col.startswith('Gene_')]
    genes = [row[col] for col in gene_columns if pd.notna(row[col])]

    # Write gene data to the sheet by checking if each gene is in DEG
    row_index = 2  # Start after header row
    for gene in genes:
        if gene in DEG['Name'].values:
            matching_rows = DEG.loc[DEG['Name'] == gene].values.tolist()
            for match_row in matching_rows:
                # Write matching row data from DEG
                for col_idx, value in enumerate(match_row, start=1):
                    sheet.cell(row=row_index, column=col_idx, value=value)
                row_index += 1

# Remove the default sheet created by Workbook() if it exists
if 'Sheet' in workbook.sheetnames:
    workbook.remove(workbook['Sheet'])

# Save the workbook to a file
workbook.save(os.path.join(output_dir, f"{input_prefix}_final_physiological_genes_in_DEG_top10pathways.xlsx"))


# Create a new Excel workbook
workbook = Workbook()

# Get column headers from the DEG DataFrame
deg_columns = list(DEG.columns)

# Iterate through the rows of 'final_disease'
for index, row in final_molecular.iterrows():
    pathway = row['Pathway']

    # Create a new sheet for each pathway with the title
    sheet = workbook.create_sheet(title=pathway[:31])  # Limit to 31 characters (Excel sheet name limit)

    # Write DEG column headers to the first row
    for col_idx, header in enumerate(deg_columns, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    # Get the gene columns from 'final_disease' row
    gene_columns = [col for col in final_molecular.columns if col.startswith('Gene_')]
    genes = [row[col] for col in gene_columns if pd.notna(row[col])]

    # Write gene data to the sheet by checking if each gene is in DEG
    row_index = 2  # Start after header row
    for gene in genes:
        if gene in DEG['Name'].values:
            matching_rows = DEG.loc[DEG['Name'] == gene].values.tolist()
            for match_row in matching_rows:
                # Write matching row data from DEG
                for col_idx, value in enumerate(match_row, start=1):
                    sheet.cell(row=row_index, column=col_idx, value=value)
                row_index += 1

# Remove the default sheet created by Workbook() if it exists
if 'Sheet' in workbook.sheetnames:
    workbook.remove(workbook['Sheet'])

# Save the workbook to a file
workbook.save(os.path.join(output_dir, f"{input_prefix}_final_molecular_genes_in_DEG_top10pathways.xlsx"))



def extract_all_subcategories(df, pathway_list):
    """Extracts all subcategories with the smallest 'B-H p-value' for a given DataFrame and pathway list.

    Args:
        df: The DataFrame containing the data.
        pathway_list: A list of pathway categories to consider.

    Returns:
        A list of all subcategories with their smallest 'B-H p-value'.
    """
    subcategory_cols = [col for col in df.columns if col.startswith('subcategory_')]
    subcategory_min_pvalues = {}

    for subcategory_col in subcategory_cols:
        # Filter the DataFrame for rows where the subcategory is in the given pathway list
        df_filtered = df[df[subcategory_col].isin(pathway_list)]
        if not df_filtered.empty:
            # Sort the filtered DataFrame by 'B-H p-value' in ascending order
            df_sorted = df_filtered.sort_values('B-H p-value')
            # Iterate over the rows to capture the minimum p-value for each unique subcategory
            for subcategory, p_value in zip(df_sorted[subcategory_col], df_sorted['B-H p-value']):
                if subcategory not in subcategory_min_pvalues or p_value < subcategory_min_pvalues[subcategory]:
                    subcategory_min_pvalues[subcategory] = p_value

    # Sort subcategories by the smallest p-value
    sorted_subcategories = sorted(subcategory_min_pvalues.items(), key=lambda item: item[1])
    all_subcategories = [subcategory for subcategory, _ in sorted_subcategories]

    return all_subcategories

disease_all = extract_all_subcategories(df_diseases, disease_list)
physiological_all = extract_all_subcategories(df_physiological, physiological_list)
molecular_all = extract_all_subcategories(df_molecular, molecular_list)

# Display the results
print("All Disease Subcategories:", disease_all)
print("All Physiological Subcategories:", physiological_all)
print("All Molecular Subcategories:", molecular_all)


def create_dfs_for_subcategories(df, subcategory_list):
  """Creates separate DataFrames for each subcategory in the list."""
  dfs = {}
  for subcategory in subcategory_list:
    dfs[subcategory] = df[df.apply(lambda row: subcategory in row.values, axis=1)]
  return dfs

# You can access individual DataFrames using the subcategory name as the key
# For example, to access the DataFrame for 'Cancer' (if it's in top_10_diseases):
#df_cardio = all_dfs['Cardiovascular System Development and Function']

dfs_diseases_all = create_dfs_for_subcategories(df_diseases, disease_all)
dfs_physiological_all = create_dfs_for_subcategories(df_physiological, physiological_all)
dfs_molecular_all = create_dfs_for_subcategories(df_molecular, molecular_all)

# Combine the dictionaries into one
All_dfs = {**dfs_diseases_all, **dfs_physiological_all, **dfs_molecular_all}

for subcategory in disease_all:
    if subcategory in All_dfs:
        df_name = "df_" + subcategory.replace(" ", "_")
        globals()[df_name] = All_dfs[subcategory]

for subcategory in physiological_all:
    if subcategory in All_dfs:
        df_name = "df_" + subcategory.replace(" ", "_")
        globals()[df_name] = All_dfs[subcategory]

for subcategory in molecular_all:
    if subcategory in All_dfs:
        df_name = "df_" + subcategory.replace(" ", "_")
        globals()[df_name] = All_dfs[subcategory]
        
        
#Print the contents of each list to verify the subcategory names
print("All Subcategories (Diseases):", disease_all)
print("All Subcategories (Physiological):", physiological_all)
print("All Subcategories (Molecular):", molecular_all)

# Print all keys in all_dfs to confirm they match the subcategory names in the lists
# prompt: list all the above df. names created

df_names = [name for name in globals() if name.startswith("df_") and isinstance(globals()[name], pd.DataFrame)]
df_names, len(df_names)

# Assuming you have df_names defined as a list of DataFrame names

# Create a new Excel workbook
workbook = Workbook()

# Iterate through DataFrame names
for df_name in df_names:
    if df_name not in ['df_diseases', 'df_physiological', 'df_molecular']:
        try:
            # Get the DataFrame from globals()
            df = globals()[df_name]

            # Create a new sheet in the workbook with the DataFrame name (truncate if necessary)
            sheet_name = df_name[:31]  # Excel sheet names are limited to 31 characters
            sheet = workbook.create_sheet(title=sheet_name)

            # Write column headers
            for col_idx, column_name in enumerate(df.columns, start=1):
                sheet.cell(row=1, column=col_idx, value=column_name)

            # Write DataFrame values, converting scientific notation to strings
            for row_idx, row in enumerate(df.values, start=2):
                for col_idx, value in enumerate(row, start=1):
                    # Convert scientific notation to string if it's a large or small float
                    if isinstance(value, float) and (value > 1e10 or value < 1e-10):
                        sheet.cell(row=row_idx, column=col_idx, value=str(value))
                    else:
                        sheet.cell(row=row_idx, column=col_idx, value=value)

        except KeyError:
            # Handle cases where a DataFrame with the given name doesn't exist
            print(f"DataFrame '{df_name}' not found.")

# Remove the default sheet created by Workbook()
if 'Sheet' in workbook.sheetnames:
    workbook.remove(workbook['Sheet'])

# Save the workbook to a file
workbook.save(os.path.join(output_dir, f"globals_dfs_subsheets_{input_prefix}_all_pathways.xlsx"))


def get_gene_list(df):
    """Extracts a master gene list from a DataFrame's 'Molecules' column."""
    if 'Molecules' not in df.columns:
        return []  # Handle cases where the column doesn't exist

    gene_list = set()  # Use a set to avoid duplicates
    for genes in df['Molecules'].dropna():  # Ignore rows with NaN in 'Molecules'
        for gene in genes.split(','):
            gene_list.add(gene.strip())  # Remove leading/trailing whitespace
    return list(gene_list)


# List of DataFrame names in globals() that start with "df_" but are not the main category DataFrames
df_names = [name for name in globals() if name.startswith("df_")
            and name not in ('df_diseases', 'df_physiological', 'df_molecular')
            and isinstance(globals()[name], pd.DataFrame)]

# List to store each DataFrame's information as a row for the final CSV
all_data = []

# Loop through each DataFrame, extract information, and add it to the list
for df_name in df_names:
    df = globals()[df_name]

    # Extract the pathway name from the DataFrame name
    pathway = df_name.replace("df_", "").replace("_", " ")

    # Determine the subcategory type based on the top 10 lists
    if pathway in disease_all:
        subcategory_type = "Diseases"
    elif pathway in physiological_all:
        subcategory_type = "Physiological"
    elif pathway in molecular_all:
        subcategory_type = "Molecular"
    else:
        subcategory_type = "Other"

    # Get the smallest "B-H p-value" in the DataFrame
    smallest_p_value = df['B-H p-value'].min() if 'B-H p-value' in df.columns else None

    # Get the master gene list
    gene_list = get_gene_list(df)
    gene_count = len(gene_list)  # Count the number of unique genes

    # Prepare the row for the CSV file
    row = {
        'Pathway': pathway,
        'Subcategory Type': subcategory_type,  # Add the new subcategory type column
        'Smallest B-H p-value': smallest_p_value,
        'Number of Genes': gene_count
    }

    # Add genes as separate columns (starting from column 5)
    for i, gene in enumerate(gene_list, start=1):
        row[f'Gene_{i}'] = gene

    # Append the row dictionary to all_data
    all_data.append(row)

# Convert all_data to a DataFrame and save to a single CSV file
final_df = pd.DataFrame(all_data)
final_df.to_csv(os.path.join(output_dir, f"all_dfs_summary_{input_prefix}_all_pathways.csv"), index=False)


